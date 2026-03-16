"""
Talk IA - Evaluateur Maileva Docs e-Facture
Usage interne - avec securite, nom obligatoire, export Excel par evaluateur

Lancement local :
  py -m pip install streamlit requests openpyxl
  streamlit run app.py

Deploiement Streamlit Cloud :
  1. Push app.py + requirements.txt sur GitHub
  2. https://share.streamlit.io -> New app
  3. Dans Settings -> Secrets, ajouter :
       TALK_API_KEY = "ta_cle_api"
       APP_PASSWORD = "le_mot_de_passe"
"""

import datetime
import io
import time
from collections import defaultdict

import requests
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
#  CONFIGURATION
#  En local : modifie directement les valeurs ci-dessous
#  En prod   : utilise st.secrets (Streamlit Cloud -> Settings -> Secrets)
# ==============================================================================

try:
    # Mode production Streamlit Cloud
    API_KEY      = st.secrets["TALK_API_KEY"]
    APP_PASSWORD = st.secrets["APP_PASSWORD"]
except Exception:
    # Mode local : mets tes valeurs ici
    API_KEY      = "COLLE_TA_CLE_ICI"
    APP_PASSWORD = "maileva2026"      # <- mot de passe pour acceder a l'app

TALK_URL = "https://talk.innovation.docaposte.com/api/TALK/ask"

# ==============================================================================
#  CONFIG PAGE
# ==============================================================================

st.set_page_config(
    page_title="Talk IA - Evaluateur Maileva",
    page_icon="💬",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ==============================================================================
#  STYLES
# ==============================================================================

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&display=swap');
  html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

  .main-header {
    background: linear-gradient(135deg, #0018A8 0%, #1a35c4 100%);
    color: white;
    padding: 28px 32px;
    border-radius: 16px;
    margin-bottom: 28px;
    box-shadow: 0 8px 32px rgba(0,24,168,0.18);
  }
  .main-header h1 { font-size: 1.7rem; font-weight: 600; margin-bottom: 4px; }
  .main-header p  { font-size: 0.95rem; opacity: 0.85; }

  .badge {
    display: inline-block;
    background: #00C8A0;
    color: #0a0f2e;
    font-size: 0.75rem;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 20px;
    margin-bottom: 12px;
  }

  .user-banner {
    background: #e8ecff;
    border: 2px solid #0018A8;
    border-radius: 12px;
    padding: 12px 20px;
    margin-bottom: 20px;
    font-size: 0.95rem;
    color: #0018A8;
    font-weight: 500;
  }

  .bubble-q {
    background: #e8ecff;
    border-left: 4px solid #0018A8;
    padding: 14px 18px;
    border-radius: 0 12px 12px 0;
    margin: 10px 0;
    font-size: 0.97rem;
    color: #0018A8;
    font-weight: 500;
  }
  .bubble-r {
    background: #f0faf7;
    border-left: 4px solid #00C8A0;
    padding: 14px 18px;
    border-radius: 0 12px 12px 0;
    margin: 10px 0;
    font-size: 0.95rem;
    line-height: 1.65;
    color: #0a0f2e;
    white-space: pre-wrap;
  }
  .card {
    background: white;
    border-radius: 14px;
    padding: 18px 22px;
    margin-bottom: 14px;
    box-shadow: 0 2px 12px rgba(0,24,168,0.07);
    border: 1px solid #e8ecff;
  }
  .card-meta { font-size: 0.78rem; color: #6b7280; margin-bottom: 8px; }
  .sbadge {
    display: inline-block;
    font-size: 0.82rem;
    font-weight: 600;
    padding: 3px 12px;
    border-radius: 20px;
    margin-top: 8px;
  }
  .s5, .s4 { background: #d1fae5; color: #065f46; }
  .s3       { background: #fef3c7; color: #92400e; }
  .s2, .s1  { background: #fee2e2; color: #991b1b; }

  .stat-row { display: flex; gap: 14px; margin-bottom: 22px; flex-wrap: wrap; }
  .stat-box {
    background: white;
    border-radius: 12px;
    padding: 14px 18px;
    flex: 1;
    min-width: 110px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,24,168,0.06);
    border: 1px solid #e8ecff;
  }
  .stat-box .val { font-size: 1.8rem; font-weight: 700; color: #0018A8; line-height: 1; }
  .stat-box .lbl { font-size: 0.72rem; color: #6b7280; margin-top: 4px; }

  .login-box {
    background: white;
    border-radius: 16px;
    padding: 36px 40px;
    box-shadow: 0 4px 24px rgba(0,24,168,0.10);
    border: 1px solid #e8ecff;
    max-width: 420px;
    margin: 60px auto;
  }
  .login-box h2 { color: #0018A8; margin-bottom: 6px; font-size: 1.3rem; }
  .login-box p  { color: #6b7280; font-size: 0.9rem; margin-bottom: 24px; }

  hr.soft { border: none; border-top: 1px solid #e8ecff; margin: 22px 0; }
</style>
""", unsafe_allow_html=True)

# ==============================================================================
#  SESSION STATE
# ==============================================================================

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "user_nom" not in st.session_state:
    st.session_state.user_nom = ""
if "user_equipe" not in st.session_state:
    st.session_state.user_equipe = ""
if "history" not in st.session_state:
    st.session_state.history = []
if "pending" not in st.session_state:
    st.session_state.pending = None

# ==============================================================================
#  ETAPE 1 : MOT DE PASSE
# ==============================================================================

if not st.session_state.authenticated:

    st.markdown("""
    <div class="main-header">
      <div class="badge">USAGE INTERNE MAILEVA</div>
      <h1>💬 Talk IA — Evaluateur</h1>
      <p>Testez Talk IA sur Maileva Docs e-Facture et notez les reponses.</p>
    </div>
    """, unsafe_allow_html=True)

    with st.container():
        st.markdown("### Acces securise")
        pwd = st.text_input(
            "Mot de passe",
            type="password",
            placeholder="Saisir le mot de passe...",
        )
        if st.button("Entrer", type="primary"):
            if pwd == APP_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Mot de passe incorrect. Contactez Amir pour obtenir l'acces.")

    st.stop()

# ==============================================================================
#  ETAPE 2 : NOM OBLIGATOIRE
# ==============================================================================

if not st.session_state.user_nom:

    st.markdown("""
    <div class="main-header">
      <div class="badge">USAGE INTERNE MAILEVA</div>
      <h1>💬 Talk IA — Evaluateur</h1>
      <p>Avant de commencer, identifiez-vous.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### Qui etes-vous ?")
    st.info("Votre nom apparaitra dans le rapport Excel pour que les resultats soient attribues correctement.")

    col1, col2 = st.columns(2)
    with col1:
        nom_input = st.text_input(
            "Prenom et Nom *",
            placeholder="ex : Sophie Martin",
        )
    with col2:
        equipe_input = st.selectbox("Equipe *", [
            "-- Selectionnez --",
            "Commercial",
            "Avant-vente",
            "Marketing",
            "Support",
            "Produit",
            "Direction",
            "Autre",
        ])

    if st.button("Commencer l'evaluation", type="primary"):
        if not nom_input.strip():
            st.error("Votre nom est obligatoire.")
        elif equipe_input == "-- Selectionnez --":
            st.error("Selectionnez votre equipe.")
        else:
            st.session_state.user_nom    = nom_input.strip()
            st.session_state.user_equipe = equipe_input
            st.rerun()

    st.stop()

# ==============================================================================
#  APPLICATION PRINCIPALE
# ==============================================================================

nom    = st.session_state.user_nom
equipe = st.session_state.user_equipe

# Header
st.markdown(f"""
<div class="main-header">
  <div class="badge">USAGE INTERNE MAILEVA</div>
  <h1>💬 Talk IA — Evaluateur</h1>
  <p>Testez Talk IA sur Maileva Docs e-Facture, notez les reponses et exportez les resultats.</p>
</div>
<div class="user-banner">
  Connecte en tant que : <strong>{nom}</strong> &nbsp;·&nbsp; Equipe : <strong>{equipe}</strong>
</div>
""", unsafe_allow_html=True)

# ==============================================================================
#  STATISTIQUES
# ==============================================================================

if st.session_state.history:
    scores = [r["score"] for r in st.session_state.history]
    avg    = round(sum(scores) / len(scores), 1) if scores else 0
    nb_ok  = sum(1 for s in scores if s >= 4)
    nb_nok = sum(1 for s in scores if s <= 2)

    st.markdown(f"""
    <div class="stat-row">
      <div class="stat-box">
        <div class="val">{len(scores)}</div>
        <div class="lbl">Questions testees</div>
      </div>
      <div class="stat-box">
        <div class="val">{avg}/5</div>
        <div class="lbl">Score moyen</div>
      </div>
      <div class="stat-box">
        <div class="val" style="color:#10b981">{nb_ok}</div>
        <div class="lbl">Bonnes (>= 4)</div>
      </div>
      <div class="stat-box">
        <div class="val" style="color:#ef4444">{nb_nok}</div>
        <div class="lbl">Mauvaises (<= 2)</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

# ==============================================================================
#  ZONE DE QUESTION
# ==============================================================================

st.markdown("<hr class='soft'>", unsafe_allow_html=True)
st.markdown("### Pose une question a Talk IA")

question_input = st.text_input(
    "Question",
    placeholder="ex : Quelle offre pour un client qui envoie 500 factures par mois ?",
    label_visibility="collapsed",
    key="question_input",
)

col_btn, _ = st.columns([2, 5])
with col_btn:
    envoyer = st.button("Envoyer a Talk IA", type="primary", use_container_width=True)

# ── Appel Talk IA ─────────────────────────────────────────────────────────────

if envoyer:
    if not question_input.strip():
        st.warning("Ecrivez une question avant d'envoyer.")
    elif st.session_state.pending is not None:
        st.warning("Notez d'abord la reponse precedente avant d'en poser une nouvelle.")
    else:
        with st.spinner("Talk IA reflechit..."):
            try:
                start = time.time()
                resp  = requests.post(
                    TALK_URL,
                    headers={
                        "Content-Type":   "application/json",
                        "x-user-api-key": API_KEY,
                    },
                    json={"prompt": question_input.strip()},
                    timeout=30,
                )
                latency = int((time.time() - start) * 1000)

                if resp.status_code == 200:
                    raw    = resp.json()
                    answer = (
                        raw.get("answer")
                        or raw.get("response")
                        or raw.get("result")
                        or raw.get("message")
                        or raw.get("text")
                        or raw.get("content")
                        or str(raw)
                    )
                    st.session_state.pending = {
                        "question":   question_input.strip(),
                        "answer":     answer,
                        "latency_ms": latency,
                        "timestamp":  datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                        "nom":        nom,
                        "equipe":     equipe,
                    }
                    st.rerun()

                elif resp.status_code == 429:
                    st.error("Rate limit atteint - attendez quelques secondes et reessayez.")
                else:
                    st.error(f"Erreur HTTP {resp.status_code} : {resp.text[:200]}")

            except requests.exceptions.Timeout:
                st.error("Talk IA n'a pas repondu dans les delais. Reessayez.")
            except Exception as e:
                st.error(f"Erreur : {e}")

# ==============================================================================
#  NOTATION
# ==============================================================================

if st.session_state.pending:
    p = st.session_state.pending

    st.markdown("<hr class='soft'>", unsafe_allow_html=True)
    st.markdown("### Reponse de Talk IA")

    st.markdown(f"""
    <div class="bubble-q">Q : {p['question']}</div>
    <div class="bubble-r">{p['answer']}</div>
    """, unsafe_allow_html=True)

    st.caption(f"Latence : {p['latency_ms']} ms")

    st.markdown("<hr class='soft'>", unsafe_allow_html=True)
    st.markdown("### Notez cette reponse")

    score = st.select_slider(
        "Votre note",
        options=[1, 2, 3, 4, 5],
        value=3,
        format_func=lambda x: {
            1: "1 — Completement faux ou hors sujet",
            2: "2 — Insuffisant ou incorrect",
            3: "3 — Correct mais incomplet",
            4: "4 — Bonne reponse, complete",
            5: "5 — Excellente reponse, precise et complete",
        }[x],
    )

    explication = st.text_area(
        "Pourquoi cette note ? (tres utile pour ameliorer le chatbot)",
        placeholder="ex : La reponse ne mentionne pas le tarif unitaire au-dela du forfait...",
        height=100,
    )

    col_val, col_ann = st.columns([2, 1])
    with col_val:
        valider = st.button("Valider la note", type="primary", use_container_width=True)
    with col_ann:
        annuler = st.button("Ignorer cette reponse", use_container_width=True)

    if valider:
        st.session_state.history.append({
            **p,
            "score":       score,
            "explication": explication.strip(),
        })
        st.session_state.pending = None
        st.success(f"Note {score}/5 enregistree !")
        st.rerun()

    if annuler:
        st.session_state.pending = None
        st.rerun()

# ==============================================================================
#  HISTORIQUE + EXPORT
# ==============================================================================

if st.session_state.history:
    st.markdown("<hr class='soft'>", unsafe_allow_html=True)

    col_titre, col_export = st.columns([3, 2])
    with col_titre:
        st.markdown(f"### Mes evaluations ({len(st.session_state.history)})")
    with col_export:
        excel_data = _build_excel(st.session_state.history)
        st.download_button(
            label="Telecharger Excel",
            data=excel_data,
            file_name=f"evaluations_talk_ia_{nom.replace(' ','_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    for entry in reversed(st.session_state.history):
        s     = entry["score"]
        label = {5:"Excellente",4:"Bonne",3:"Correcte",2:"Insuffisante",1:"Mauvaise"}.get(s,"")
        stars = "★" * s + "☆" * (5 - s)
        expl  = (f"<br><span style='font-size:0.85rem;color:#6b7280;'>"
                 f"Commentaire : {entry['explication']}</span>"
                 if entry.get("explication") else "")

        st.markdown(f"""
        <div class="card">
          <div class="card-meta">
            {entry['timestamp']} &nbsp;·&nbsp;
            {entry['nom']} ({entry['equipe']}) &nbsp;·&nbsp;
            {entry['latency_ms']} ms
          </div>
          <div class="bubble-q">Q : {entry['question']}</div>
          <div style="font-size:0.9rem;color:#374151;line-height:1.6;
                      padding:6px 0;white-space:pre-wrap;">
            {entry['answer'][:500]}{'...' if len(entry['answer'])>500 else ''}
          </div>
          <span class="sbadge s{s}">{stars} {label}</span>
          {expl}
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<hr class='soft'>", unsafe_allow_html=True)
    if st.button("Reinitialiser ma session"):
        st.session_state.history = []
        st.session_state.pending = None
        st.rerun()

# ==============================================================================
#  GENERATION EXCEL
# ==============================================================================

def _build_excel(history: list) -> bytes:

    C_BLUE  = "0018A8"
    C_TEAL  = "00C8A0"
    C_WHITE = "FFFFFF"
    C_DARK  = "002080"
    C_GREEN = "D5F5E3"
    C_ORNG  = "FEF9E7"
    C_RED   = "FADBD8"

    def _f(h):   return PatternFill("solid", fgColor=h)
    def _ft(bold=False, size=10, color="000000"):
        return Font(name="Calibri", size=size, bold=bold, color=color)
    def _a(h="left", v="center", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def _b():
        s = Side(border_style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)
    def _sbg(s):
        return C_GREEN if s >= 4 else (C_ORNG if s == 3 else C_RED)

    wb = openpyxl.Workbook()

    # ── Onglet 1 : Toutes les evaluations ─────────────────────────────────────
    ws = wb.active
    ws.title = "Toutes les evaluations"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "EVALUATIONS TALK IA - Maileva Docs e-Facture"
    c.font      = _ft(bold=True, size=14, color=C_WHITE)
    c.fill      = _f(C_BLUE)
    c.alignment = _a("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value     = (f"Exporte le {datetime.datetime.now().strftime('%d/%m/%Y a %H:%M')}"
                    f"   -   {len(history)} evaluations")
    c2.font      = _ft(size=10, color=C_WHITE)
    c2.fill      = _f(C_DARK)
    c2.alignment = _a("center")
    ws.row_dimensions[2].height = 18

    COLS = [
        ("#",            4),
        ("Date/Heure",  18),
        ("Evaluateur",  22),
        ("Equipe",      15),
        ("Question",    52),
        ("Reponse",     72),
        ("Score /5",     9),
        ("Commentaire", 52),
    ]
    for col, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = _ft(bold=True, size=11, color=C_WHITE)
        c.fill      = _f(C_TEAL)
        c.alignment = _a("center", wrap=False)
        c.border    = _b()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 26

    for i, r in enumerate(history, 1):
        row   = i + 3
        score = r.get("score", 0)
        bg    = "F9FAFB" if i % 2 == 0 else C_WHITE

        for col, val in enumerate([
            i,
            r.get("timestamp", ""),
            r.get("nom", ""),
            r.get("equipe", ""),
            r.get("question", ""),
            r.get("answer", ""),
            score,
            r.get("explication", ""),
        ], 1):
            c        = ws.cell(row=row, column=col, value=val)
            c.border = _b()
            if col == 7:
                c.fill      = _f(_sbg(score))
                c.alignment = _a("center", wrap=False)
                c.font      = _ft(bold=True, size=13)
            else:
                c.fill      = _f(bg)
                c.alignment = _a("left", "top", True)
                c.font      = _ft(size=10)
        ws.row_dimensions[row].height = 80

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{len(history)+3}"

    # ── Onglet 2 : Stats par evaluateur ───────────────────────────────────────
    ws2 = wb.create_sheet("Stats par evaluateur")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    t = ws2["A1"]
    t.value     = "Performances par evaluateur"
    t.font      = _ft(bold=True, size=13, color=C_WHITE)
    t.fill      = _f(C_BLUE)
    t.alignment = _a("center")
    ws2.row_dimensions[1].height = 28

    COLS2 = [
        ("Evaluateur",     24),
        ("Equipe",         16),
        ("Nb questions",   14),
        ("Score moyen /5", 16),
        ("% Bonnes (>=4)", 16),
        ("% Mauvaises (<=2)", 18),
    ]
    for col, (h, w) in enumerate(COLS2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font      = _ft(bold=True, size=11, color=C_WHITE)
        c.fill      = _f(C_TEAL)
        c.alignment = _a("center", wrap=False)
        c.border    = _b()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 24

    # Grouper par evaluateur
    par_eval = defaultdict(list)
    for r in history:
        key = (r.get("nom", "Anonyme"), r.get("equipe", ""))
        par_eval[key].append(r.get("score", 0))

    row_n = 3
    for (eval_nom, eval_eq), eval_scores in sorted(par_eval.items()):
        n      = len(eval_scores) or 1
        avg_s  = round(sum(eval_scores) / n, 2)
        pct_ok = f"{round(sum(1 for s in eval_scores if s >= 4)/n*100, 1)} %"
        pct_ko = f"{round(sum(1 for s in eval_scores if s <= 2)/n*100, 1)} %"
        bg     = _sbg(avg_s)

        for col, val in enumerate([eval_nom, eval_eq, n, avg_s, pct_ok, pct_ko], 1):
            c        = ws2.cell(row=row_n, column=col, value=val)
            c.font   = _ft(bold=(col in (4, 5)), size=11)
            c.border = _b()
            c.fill   = _f(bg if col in (4, 5) else ("F9FAFB" if row_n % 2 == 0 else C_WHITE))
            c.alignment = _a("center" if col > 2 else "left", wrap=False)
        row_n += 1

    ws2.freeze_panes = "A3"

    # ── Onglet 3 : Stats globales ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Stats globales")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 22

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value     = "Statistiques globales"
    t3.font      = _ft(bold=True, size=13, color=C_WHITE)
    t3.fill      = _f(C_BLUE)
    t3.alignment = _a("center")
    ws3.row_dimensions[1].height = 28

    all_scores = [r.get("score", 0) for r in history]
    n          = len(all_scores) or 1

    for row_i, (label, val) in enumerate([
        ("Total evaluations",           len(history)),
        ("Nombre d'evaluateurs",        len(par_eval)),
        ("Score moyen global (/5)",     round(sum(all_scores)/n, 2)),
        ("% Excellentes (5/5)",         f"{round(sum(1 for s in all_scores if s==5)/n*100,1)} %"),
        ("% Bonnes (4/5)",              f"{round(sum(1 for s in all_scores if s==4)/n*100,1)} %"),
        ("% Correctes (3/5)",           f"{round(sum(1 for s in all_scores if s==3)/n*100,1)} %"),
        ("% Insuffisantes (2/5)",       f"{round(sum(1 for s in all_scores if s==2)/n*100,1)} %"),
        ("% Mauvaises (1/5)",           f"{round(sum(1 for s in all_scores if s==1)/n*100,1)} %"),
        ("Date export",                 datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
    ], 2):
        bg = "F3F4F6" if row_i % 2 == 0 else C_WHITE
        cl = ws3.cell(row=row_i, column=1, value=label)
        cv = ws3.cell(row=row_i, column=2, value=val)
        for c in [cl, cv]:
            c.fill      = _f(bg)
            c.alignment = _a("left", wrap=False)
        cl.font = _ft(size=11)
        cv.font = _ft(bold=True, size=11, color=C_BLUE)
        ws3.row_dimensions[row_i].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

